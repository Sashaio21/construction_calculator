import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment



# Функция для перевода ключей словаря на русский
def translate_keys(data):
    translations = {
        'name': 'Имя',
        'numberStoreys': 'Количество этажей',
        'square': 'Площадь',
        'mansard': 'Мансарда',
        'WallMaterial': 'Материал стен',
        'RoofMaterial': 'Материал крыши',
        'FoundationMaterial': 'Материал фундамента',
        'FacadeMaterial': 'Материал фасада',
        'number': 'Номер'
    }
    translated_data = {translations.get(key, key): value for key, value in data.items()}
    return translated_data

def createExcelFile(data):
    # Переводим ключи на русский
    translated_data = translate_keys(data)

    # Преобразуем данные в DataFrame
    df = pd.DataFrame([translated_data])

    # Сохраняем DataFrame в Excel файл
    file_name = 'building_info.xlsx'
    df.to_excel(file_name, index=False)

    # Загружаем файл для изменения стилей
    wb = load_workbook(file_name)
    ws = wb.active

    # Применяем стиль к заголовкам
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Устанавливаем перенос строк для всех ячеек
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)


    # Устанавливаем ширину столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    wb.save("building_info.xlsx")


# data = {
#         'name': 'Дcdscdscdsccdsима',
#         'numberStoreys': '3',
#         'square': '60',
#         'mansard': 'Да',
#         'WallMaterial': 'Кладка бессер блоков',
#         'RoofMaterial': 'Кровля из композитной черепицы',
#         'FoundationMaterial': 'Заливка свайно-ростверкового фундамента',
#         'FacadeMaterial': 'Обшивка вагонкой',
#         'number': '375292119326'
#     }

# wb = createExcelFile(data)

# # Сохраняем изменения в Excel файл
# wb.save("building_info.xlsx")

# print("Excel файл успешно создан: building_info.xlsx")
